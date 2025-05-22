from io import BytesIO
import re
import tempfile
from openpyxl import load_workbook
import streamlit as st
import pandas as pd
import os
import base64
from jinja2 import Template
import plotly.graph_objects as go
import plotly.io as pio
from fpdf import FPDF
import glob
from constants.constants import (
    KEY_LISTING_NAME,
    KEY_REVPAR_INDEX,
    KEY_REVPAR_INDEX_STLY,
    KEY_REVPAR_STLY,
    KEY_MARKET_REVPAR_STLY,
    KEY_MARKET_PEN,
    KEY_MARKET_PEN_STLY,
    KEY_PAID_OCCUPANCY_STLY,
    KEY_MARKET_OCCUPANCY_STLY,
    KEY_ADR_INDEX_STLY,
    KEY_ADR_STLY,
    KEY_MARKET_ADR_STLY,
    KEY_LABELS,
    customers
)


def get_diff_percent_bar(df: pd.DataFrame, x: str, y: str, title: str, yaxis_title: str, base: int):
    df = df.sort_values(by=[y], ignore_index=True)
    x_vals = df[x]
    y_vals = df[y] - base  # offset from base
    base_vals = [1] * len(df)

    # Build the custom bar chart
    fig = go.Figure()

    fig.add_trace(go.Bar(
        x=x_vals,
        y=y_vals,
        base=base_vals,
        marker_color=["green" if y > 0 else ("red" if y > -1 else "gray") for y in y_vals],
        hovertext=df[KEY_LISTING_NAME],
        hoverinfo="text+y"
    ))

    fig.update_layout(
        yaxis_title=yaxis_title,
        yaxis=dict(
            zeroline=False,
            showgrid=True
        ),
        title=title,
        shapes=[
            dict(
                type="line",
                x0=-0.5,
                x1=len(df) - 0.5,
                y0=base,
                y1=base,
                line=dict(color="black", dash="dash", width=1)
            )
        ],
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        font=dict(color="white")
    )
    return fig

def charts_for_listing(row):
    def make_comparison_chart(title, current_val, stly_val):
        fig = go.Figure()
        fig.add_trace(go.Bar(
            x=["Current"],
            y=[current_val],
            name="Current",
            marker_color="lightgray",
            text=[f"{current_val:.0%}"],
            textposition="auto",
            textfont=dict(size=20)
        ))
        fig.add_trace(go.Bar(
            x=["STLY"],
            y=[stly_val],
            name="STLY",
            marker_color="black",
            text=[f"{stly_val:.0%}"],
            textposition="auto",
            textfont=dict(size=20)
        ))
        fig.update_layout(
            title=dict(
                text=title,
                font=dict(size=30),
                x=0.5,  # Center title
                xanchor='center'
            ),
            barmode="group",
            xaxis=dict(showticklabels=False),  # Removes x-axis labels
            legend=dict(
                orientation="h",
                yanchor="bottom",
                y=0.96,
                xanchor="right",
                x=1
            ),
            font=dict(color="black"),
            margin=dict(t=50)
        )
        return fig

    return [
        make_comparison_chart("RevPAR Index", row[KEY_REVPAR_INDEX], row[KEY_REVPAR_INDEX_STLY]),
        make_comparison_chart("Market Penetration Index", row[KEY_MARKET_PEN], row[KEY_MARKET_PEN_STLY]),
    ]


def render_upload_page():
    # Set the title that appears at the top of the page.
    st.image("images/enrich_logo.png")
    '''
    # :earth_americas: Enrich Revenue Dashboard
    '''

    # Add some spacing
    ''
    ''
    # Add a form for customer selection and file upload
    with st.form(key="customer_file_form"):
        selected_customer = st.selectbox(
            'Which customer is this report for?',
            customers,
            help="Select a customer from the dropdown."
        )
        # Dropdowns for selecting month and year
        month = st.selectbox(
            "Select Month",
            options=[
                "January", "February", "March", "April", "May", "June",
                "July", "August", "September", "October", "November", "December"
            ]
        )
        year = st.selectbox(
            "Select Year",
            options=[str(y) for y in range(2020, pd.Timestamp.now().year + 1)]
        )
        uploaded_file = st.file_uploader(
            "Upload a file", 
            type=["csv", "xlsx"], 
            help="Upload a CSV or Excel file."
        )
        submit_button = st.form_submit_button(label="Generate Report", help="Click to generate the report.")

    # Ensure both fields are filled before proceeding
    if submit_button:
        if not selected_customer:
            st.error("Please select a customer.")
        elif not uploaded_file:
            st.error("Please upload a file.")
        else:
            #Process the uploaded file
            try:
                df = pd.read_csv(uploaded_file)
            except:
                df = pd.read_excel(uploaded_file)

            if df is not None:
                df = df.fillna('')
                logo_paths = glob.glob(os.path.join("customer_logos", selected_customer + ".*"))
                logo_path = logo_paths[0]

                # Calculations
                df[KEY_REVPAR_INDEX] = df[KEY_REVPAR_INDEX] / 100
                df[KEY_MARKET_PEN] = df[KEY_MARKET_PEN] / 100
                df[KEY_REVPAR_INDEX_STLY] = df[KEY_REVPAR_STLY] / df[KEY_MARKET_REVPAR_STLY]
                df[KEY_MARKET_PEN_STLY] = df[KEY_PAID_OCCUPANCY_STLY] / df[KEY_MARKET_OCCUPANCY_STLY]
                df[KEY_ADR_INDEX_STLY] = df[KEY_ADR_STLY] / df[KEY_MARKET_ADR_STLY]

                # Create Charts
                df[KEY_LABELS] = df[KEY_LISTING_NAME].str[:20] + "..."
                rpi_thisPeriod = get_diff_percent_bar(df, KEY_LABELS, KEY_REVPAR_INDEX, "RevPAR Index this Period", "RevPAR Index", 1)
                rpi_stly = get_diff_percent_bar(df, KEY_LABELS, KEY_REVPAR_INDEX_STLY, "RevPar Index STLY", "RevPar Index", 1)
                mpi_thisPeriod = get_diff_percent_bar(df, KEY_LABELS, KEY_MARKET_PEN, "Market Penetration Index", "MPI", 1)
                mpi_stly = get_diff_percent_bar(df, KEY_LABELS, KEY_MARKET_PEN_STLY, "Market Penetration Index STLY", "MPI", 1)
    
                def get_imgs(charts):
                    img_bytes = []
                    for chart in charts:
                        buffer = BytesIO(pio.to_image(chart, format='png'))
                        buffer.name = f"{chart['layout']['title']['text']}.png".replace(" ", "_")
                        buffer.seek(0)
                        img_bytes.append(buffer)
                    return img_bytes

                charts = [rpi_thisPeriod, rpi_stly, mpi_thisPeriod, mpi_stly]
                imgs = get_imgs(charts)
                # Create a PDF document
                pdf = FPDF()
                pdf.set_auto_page_break(auto=True, margin=15)
                pdf.add_page()
                pdf.set_font("Arial", size=12)
                pdf.image(logo_path, x=10, y=8, w=33)
                pdf.cell(200, 10, txt=f"Monthly Revenue Report - {selected_customer} ", ln=True, align="C")
                pdf.ln(20)
                temp_paths = []
                for i, chart in enumerate(charts):
                    chart.update_layout(
                        font=dict(color="black")
                    )
                    with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
                        chart.write_image(tmp, format="png")
                        temp_paths.append(tmp.name)
                        pdf.image(tmp.name, x=10 + ((i%2)*90), y=pdf.get_y(), w=80)
                        if i % 2 == 1:
                            pdf.ln(90)
                
                for i, row in df.iterrows():
                    listing_name = row[KEY_LISTING_NAME]
                    pdf.add_page()
                    # Header
                    pdf.set_fill_color(200, 190, 180)  # muted beige background
                    pdf.rect(0, 0, 210, 35, style='F')
                    pdf.set_xy(10, 8)
                    pdf.set_font("Arial", "B", 20)
                    pdf.set_text_color(0, 0, 0)
                    pdf.cell(0, 10, listing_name, ln=True)
                    pdf.set_xy(10, 18)
                    pdf.set_font("Arial", "", 12)
                    print(f"{15 - i}")
                    pdf.cell(0, 10, f"{month} Report - {row['Group Name']}", ln=False)
                    # Logo
                    pdf.image(logo_path, x=175, y=5, w=25)
                    # KPI Boxes
                    pdf.set_font("Arial", "B", 12)
                    pdf.set_text_color(0, 0, 0)
                    kpi_box_start_x = 10
                    box_width = 60
                    box_margin = 5
                    kpi_box_y = 45
                    def draw_kpi(label, value, x, label_font_size):
                        pdf.set_xy(x, kpi_box_y)
                        pdf.set_font("Arial", "", label_font_size)
                        pdf.cell(box_width, 10, txt=label, ln=True, border=1, align="C")
                        pdf.set_xy(x, kpi_box_y + 10)
                        pdf.set_font("Arial", "", 16)
                        pdf.cell(box_width, 10, txt=str(value), ln=True, border=1, align="C")

                    draw_kpi("Market Penetration Index", f"{int(row[KEY_MARKET_PEN])}", kpi_box_start_x, 14)
                    # draw_kpi("Booked Nights Pickup\n(Past 30 Days)", int(row["Booked Nights Pickup (30 Days)"]), kpi_box_start_y + box_height + box_margin)
                    draw_kpi("Booked Nights Pickup\n(Past 30 Days)", 10, kpi_box_start_x + box_width + box_margin, 9)
                    draw_kpi("RevPAR Index", f"{int(row[KEY_REVPAR_INDEX])}%", kpi_box_start_x + 2*(box_width + box_margin), 15)
                    # Charts
                    chart_y_start = 75
                    chart_x_spacing = 90
                    chart_y_spacing = 60

                    for i, fig in enumerate(charts_for_listing(row)):  # You'd define this to generate or return 4 Plotly charts
                        x = 10 + (i % 2) * chart_x_spacing
                        y = chart_y_start + (i // 2) * chart_y_spacing

                        with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
                            fig.update_layout(font=dict(color="black"), title_font_size=30)
                            fig.write_image(tmp.name, format="png")
                            pdf.image(tmp.name, x=x, y=y, w=80)

                # Save the PDF to a BytesIO object
                pdf_bytes = pdf.output(dest="S").encode("latin-1")  # fpdf uses latin-1 encoding
                pdf_output = BytesIO(pdf_bytes)
                for path in temp_paths:
                    os.remove(path)
                if st.download_button(
                    label="Download Report",
                    data=pdf_output,
                    file_name="report.pdf",
                    mime="application/pdf",
                ):
                    st.success("Report downloaded successfully.")